# scripts/extract-data-task.py
# Read "data-task.xlsx" and emit grouped orders JSON for import-data-task.mjs.
#
# Sheet columns (1-based):
#   1 SR.  2 Name  3 Email  4 Phone  5 Months  6 Start Date  7 End Date  8 Order ID  9 Amount  10.. Reports Sold
#
# Name/SR/Email are only on the first row of each user block -> forward-filled.
# Rows that share an Order ID = ONE order. The total Amount sits on the first row of
# the group (continuation rows are 0). Each row contributes line items (its products +
# its own start/end/months).
#
# One Order ID -> { name, orderId, amount, purchaseDate(min start), expiryDate(max end),
#                   items:[{product,start,end,months}], issues:[...] }
#
# Bad rows (blank/unparseable dates, non-numeric amount, end<start) are flagged in
# `issues` and the order is also listed in the printed report. Nothing is invented.
#
# Usage:
#   python scripts/extract-data-task.py "<path to data-task.xlsx>" [out.json] [--only "Ross Rizzo"]

import sys, json, re, datetime, calendar

def parse_date(v):
    """Return (iso_string|None, status). status: ok|blank|<error tag>."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None, 'blank'
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d'), 'ok'
    s = str(v).strip()
    m = re.match(r'^(\d{1,2})-+(\d{1,2})-(\d{1,4})$', s)        # D-M-YYYY (incl. "11--06-2026")
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.datetime(y, mo, d).strftime('%Y-%m-%d'), 'ok'
        except ValueError:
            return None, f'baddate:{s}'
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)            # ISO
    if m:
        try:
            return datetime.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).strftime('%Y-%m-%d'), 'ok'
        except ValueError:
            return None, f'baddate:{s}'
    return None, f'unparsed:{s}'

def main():
    args = [a for a in sys.argv[1:]]
    only = None
    if '--only' in args:
        i = args.index('--only'); only = args[i + 1]; del args[i:i + 2]
    src = args[0] if args else r'D:\new Downloads\Downloads\data-task.xlsx'
    out = args[1] if len(args) > 1 else None

    import openpyxl
    ws = openpyxl.load_workbook(src, data_only=True).active

    # 1) read raw rows, forward-filling name + email
    raw = []
    name, email = None, None
    for r in range(2, ws.max_row + 1):
        cell = lambda i: ws.cell(r, i).value
        if cell(2):
            name = str(cell(2)).strip()
            email = str(cell(3)).strip() if cell(3) else ''
        oid = cell(8)
        if oid is None and all(cell(i) is None for i in range(5, 10)):
            continue  # fully blank row
        prods = [str(cell(i)).strip() for i in range(10, ws.max_column + 1)
                 if cell(i) not in (None, '') and str(cell(i)).strip()]
        raw.append(dict(row=r, name=name, email=(email or ''), months=cell(5), start=cell(6),
                        end=cell(7), oid=(str(oid).strip() if oid else None),
                        amount=cell(9), products=prods))

    if only:
        raw = [x for x in raw if (x['name'] or '').strip().lower() == only.strip().lower()]

    # 2) group by Order ID (preserve first-seen order)
    groups = {}
    order_seq = []
    for x in raw:
        oid = x['oid'] or f"__noid_row{x['row']}"
        if oid not in groups:
            groups[oid] = []
            order_seq.append(oid)
        groups[oid].append(x)

    orders = []
    for oid in order_seq:
        rows = groups[oid]
        name = rows[0]['name']
        email = rows[0].get('email', '')
        issues = []
        amount = 0.0
        for x in rows:
            a = x['amount']
            if isinstance(a, (int, float)):
                amount += a
            elif a not in (None, ''):
                issues.append(f"row {x['row']}: bad amount {a!r}")
        items, starts, ends = [], [], []
        for x in rows:
            sd, ss = parse_date(x['start'])
            ed, es = parse_date(x['end'])
            if ss != 'ok':
                issues.append(f"row {x['row']}: start {ss}")
            if es != 'ok':
                issues.append(f"row {x['row']}: end {es}")
            if sd and ed and ed < sd:
                issues.append(f"row {x['row']}: end<start ({sd} > {ed})")
            if sd:
                starts.append(sd)
            if ed:
                ends.append(ed)
            months = x['months'] if isinstance(x['months'], (int, float)) else None
            for p in x['products']:
                items.append(dict(product=p, start=sd, end=ed, months=months, srcRow=x['row']))
        orders.append(dict(
            name=name, email=email,
            orderId=(oid if not oid.startswith('__noid') else None),
            amount=round(amount, 2),
            purchaseDate=(min(starts) if starts else None),
            expiryDate=(max(ends) if ends else None),
            items=items, issues=issues,
        ))

    # 3) report
    clean = [o for o in orders if not o['issues']]
    bad = [o for o in orders if o['issues']]
    print(f"Orders: {len(orders)}  (clean {len(clean)}, flagged {len(bad)})  from {len(raw)} rows"
          + (f'  [only "{only}"]' if only else ''))
    if bad:
        print("\nFLAGGED ORDERS:")
        for o in bad:
            print(f"  {o['orderId'] or '(no id)'} | {o['name']} | {'; '.join(o['issues'])}")

    if out:
        with open(out, 'w', encoding='utf-8') as f:
            json.dump(orders, f, ensure_ascii=False, indent=2)
        print(f"\nWrote {len(orders)} orders -> {out}")

if __name__ == '__main__':
    main()
